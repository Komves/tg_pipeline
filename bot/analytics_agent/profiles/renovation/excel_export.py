from __future__ import annotations

from pathlib import Path
from typing import Any, Dict
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from analytics_agent.profiles.renovation.surface_estimator import estimate_surfaces
from analytics_agent.profiles.renovation.material_basket import build_material_basket
from analytics_agent.profiles.renovation.market_pricing import price_material_basket
from analytics_agent.profiles.renovation.labor_estimator import estimate_labor


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def build_renovation_xlsx(task_id: str, params: Dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    surfaces = estimate_surfaces(params)
    basket = build_material_basket(params, surfaces)
    priced = price_material_basket(basket)

    wb = Workbook()

    area = float(params.get("area_m2") or 0)
    estimate_scope = params.get("estimate_scope") or "materials"

    CATEGORY_TITLES = {
        "flooring": "Напольное покрытие",
        "bathroom_tile": "Плитка санузла",
        "plinth": "Плинтус",
        "primer": "Грунтовка",
        "putty": "Шпаклёвка",
        "paint_or_wallpaper": "Краска / обои",
        "rough_mix": "Ровнитель / смеси",
        "tile_adhesive": "Плиточный клей",
        "grout": "Затирка",
    }

    ws2 = wb.active
    ws2.title = "Материалы"
    ws2.append([
        "Категория",
        "Позиция",
        "Количество",
        "Ед.",
        "Цена",
        "Итого",
        "Источник",
    ])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    row_idx = 2
    for item in priced.get("items") or []:
        qty = item.get("required_packs") if item.get("pricing_mode") == "by_pack" else item.get("quantity")
        unit_price = item.get("unit_price_rub") if item.get("usable_for_total") else None

        ws2.append([
            CATEGORY_TITLES.get(item.get("category"), item.get("category")),
            item.get("name"),
            qty,
            item.get("market_unit") or item.get("unit"),
            unit_price,
            f"=C{row_idx}*E{row_idx}",
            item.get("source_title") or item.get("pricing_source"),
        ])
        row_idx += 1

    ws3 = wb.create_sheet("Работы")
    ws3.append(["Позиция", "Объем", "Ед.", "Ставка", "Итого"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    if estimate_scope == "materials_and_labor":
        labor = estimate_labor(params)

        labor_row_idx = 2

        for labor_item in labor.get("labor_items") or []:
            ws3.append([
                labor_item.get("title_ru"),
                labor_item.get("quantity"),
                labor_item.get("unit"),
                labor_item.get("unit_price_rub"),
                f"=B{labor_row_idx}*D{labor_row_idx}",
            ])
            labor_row_idx += 1

    ws_total = wb.create_sheet("Итого")

    ws_total.append(["Показатель", "Сумма"])

    for cell in ws_total[1]:
        cell.font = Font(bold=True)

    ws_total.append([
        "Материалы",
        f"=SUM(Материалы!F2:F999)",
    ])

    if estimate_scope == "materials_and_labor":
        ws_total.append([
            "Работы",
            f"=SUM(Работы!E2:E999)",
        ])

        ws_total.append([
            "Общий итог",
            "=B2+B3",
        ])

    for sheet in wb.worksheets:
        _auto_width(sheet)

    path = output_dir / f"renovation_estimate_{task_id}_{uuid4().hex[:6]}.xlsx"
    wb.save(path)
    return path